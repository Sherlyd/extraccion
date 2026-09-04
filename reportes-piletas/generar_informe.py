# generar_informe.py
# Arma un Excel descargable con los mismos datos que ve el usuario en
# el dashboard web -- ya filtrados por su rol. Se genera al vuelo cada
# vez que alguien pide la descarga, nunca se guarda un archivo viejo
# que pueda quedar desactualizado.

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
NORMAL_FONT = Font(name=FONT_NAME, size=10)


def _escribir_tabla(ws, encabezados, filas, start_row, money_cols=None):
    money_cols = money_cols or []
    for j, col in enumerate(encabezados, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')

    for i, fila in enumerate(filas, start=start_row + 1):
        for j, valor in enumerate(fila, start=1):
            c = ws.cell(row=i, column=j, value=valor)
            c.font = NORMAL_FONT
            if j - 1 in money_cols:
                c.number_format = '$#,##0'

    for j, col in enumerate(encabezados, start=1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else 'A'].width = max(14, len(col) + 2)

    return start_row + len(filas) + 2


def generar_informe_excel(usuario, por_mes, comparacion, top_articulos, top_distribuidores,
                           cartera, alertas, por_sucursal=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'

    ws['B2'] = f"Informe Piletas — {usuario['nombre']} ({usuario['rol']})"
    ws['B2'].font = TITLE_FONT
    ws['B3'] = 'Generado automaticamente desde el dashboard'
    ws['B3'].font = Font(name=FONT_NAME, size=9, italic=True, color='808080')

    row = 5
    if comparacion:
        ws.cell(row=row, column=2, value='Facturación mes actual').font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=row, column=4, value=comparacion['valor_actual']).number_format = '$#,##0'
        row += 1
        ws.cell(row=row, column=2, value='Variación vs. promedio histórico').font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=row, column=4, value=comparacion['variacion_pct']).number_format = '0.0%'
        row += 1
    ws.cell(row=row, column=2, value='Pendiente de fabricar/entregar').font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=row, column=4, value=cartera['total']).number_format = '$#,##0'
    row += 1
    ws.cell(row=row, column=2, value='Bloqueado por ONF').font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=row, column=4, value=cartera['bloqueado_onf']).number_format = '$#,##0'
    row += 2

    if alertas:
        ws.cell(row=row, column=2, value='Alertas').font = Font(name=FONT_NAME, bold=True, color='9C0006')
        row += 1
        for a in alertas:
            ws.cell(row=row, column=2, value=f'- {a}').font = Font(name=FONT_NAME, color='9C0006')
            row += 1

    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['D'].width = 18

    ws2 = wb.create_sheet('Facturación Mensual')
    _escribir_tabla(ws2, ['Mes', 'Importe Neto'], list(por_mes.items()), 1, money_cols=[1])

    ws3 = wb.create_sheet('Top Artículos')
    _escribir_tabla(ws3, ['Artículo', 'Importe'], top_articulos, 1, money_cols=[1])

    ws4 = wb.create_sheet('Top Distribuidores')
    _escribir_tabla(ws4, ['Distribuidor', 'Importe'], top_distribuidores, 1, money_cols=[1])

    if por_sucursal:
        ws5 = wb.create_sheet('Por Sucursal')
        _escribir_tabla(ws5, ['Sucursal', 'Importe'], por_sucursal, 1, money_cols=[1])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
