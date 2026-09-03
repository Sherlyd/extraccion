# generar_informe.py
# Construye el archivo Excel final con una hoja de resumen ejecutivo y
# una hoja por cada cruce, con formato profesional y un grafico de la
# evolucion mensual.

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
NORMAL_FONT = Font(name=FONT_NAME, size=10)
RISK_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
RISK_FONT = Font(name=FONT_NAME, size=10, color='9C0006')
THIN_BORDER = Border(bottom=Side(style='thin', color='D9D9D9'))


def _write_table(ws, df, start_row=1, money_cols=None, pct_cols=None, date_cols=None, highlight_col=None):
    """Escribe un DataFrame como tabla formateada a partir de start_row."""
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    date_cols = date_cols or []

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        highlight = highlight_col is not None and getattr(row, highlight_col, False)
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = RISK_FONT if highlight else NORMAL_FONT
            if highlight:
                cell.fill = RISK_FILL
            cell.border = THIN_BORDER

            col_name = df.columns[j - 1]
            if col_name in money_cols:
                cell.number_format = '$#,##0'
            elif col_name in pct_cols:
                cell.number_format = '0.0%'
            elif col_name in date_cols:
                cell.number_format = 'DD-MM-YYYY'

    for j, col in enumerate(df.columns, start=1):
        width = max(12, min(40, int(df[col].astype(str).str.len().max() or 10) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width

    return start_row + len(df) + 1  # proxima fila libre


def generar_informe(output_path, mensual, top_cli, pareto_resumen, pareto_detalle,
                     cartera_cli, cruce_riesgo, kpis):
    wb = Workbook()

    # ---------- Hoja 1: Resumen ejecutivo ----------
    ws = wb.active
    ws.title = 'Resumen'
    ws['B2'] = 'Informe comercial — cruce de datos Qlik'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = 'Generado automaticamente a partir de los exports de Qlik Sense'
    ws['B3'].font = Font(name=FONT_NAME, size=9, italic=True, color='808080')

    row = 5
    for label, value, fmt in kpis:
        ws.cell(row=row, column=2, value=label).font = Font(name=FONT_NAME, bold=True)
        c = ws.cell(row=row, column=4, value=value)
        c.font = NORMAL_FONT
        if fmt:
            c.number_format = fmt
        row += 1
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['D'].width = 20

    # ---------- Hoja 2: Evolucion mensual ----------
    ws2 = wb.create_sheet('Evolucion Mensual')
    ws2['B1'] = 'Facturacion neta por mes'
    ws2['B1'].font = TITLE_FONT
    next_row = _write_table(
        ws2, mensual, start_row=3,
        money_cols=['Importe_Neto'], pct_cols=[], date_cols=['mes'], highlight_col='Atipico',
    )

    chart = BarChart()
    chart.title = 'Facturacion neta mensual'
    chart.y_axis.title = '$'
    chart.x_axis.title = 'Mes'
    data = Reference(ws2, min_col=2, min_row=3, max_row=3 + len(mensual))
    cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(mensual))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 22, 10
    ws2.add_chart(chart, f'B{next_row + 2}')

    # ---------- Hoja 3: Top clientes ----------
    ws3 = wb.create_sheet('Top Clientes')
    ws3['B1'] = 'Ranking de clientes por facturacion neta'
    ws3['B1'].font = TITLE_FONT
    _write_table(ws3, top_cli, start_row=3, money_cols=['Importe_Neto'], pct_cols=['% del Total'])

    # ---------- Hoja 4: Concentracion de distribuidores ----------
    ws4 = wb.create_sheet('Concentracion Distrib.')
    ws4['B1'] = 'Concentracion de la facturacion (Pareto)'
    ws4['B1'].font = TITLE_FONT
    next_row = _write_table(ws4, pareto_resumen, start_row=3, pct_cols=['% de la facturacion total'])
    ws4.cell(row=next_row + 1, column=2, value='Detalle completo, ordenado de mayor a menor:').font = Font(
        name=FONT_NAME, bold=True)
    _write_table(ws4, pareto_detalle, start_row=next_row + 3,
                 money_cols=['M$ Fact.'], pct_cols=['Fact. %', '% Acumulado'])

    # ---------- Hoja 5: Cartera pendiente ----------
    ws5 = wb.create_sheet('Cartera Pendiente')
    ws5['B1'] = 'Cartera pendiente de cobro por cliente'
    ws5['B1'].font = TITLE_FONT
    _write_table(ws5, cartera_cli, start_row=3,
                 money_cols=['Total_Pendiente', 'Con_ONF'], pct_cols=['% Bloqueado por ONF'])

    # ---------- Hoja 6: Cruce de riesgo (facturacion vs cartera) ----------
    ws6 = wb.create_sheet('Riesgo Cartera')
    ws6['B1'] = 'Cruce: facturacion vs. cartera pendiente por cliente'
    ws6['B1'].font = TITLE_FONT
    ws6['B2'] = 'Filas resaltadas: cartera pendiente supera el umbral definido en config.py'
    ws6['B2'].font = Font(name=FONT_NAME, size=9, italic=True, color='808080')
    _write_table(ws6, cruce_riesgo, start_row=4,
                 money_cols=['Facturacion', 'Cartera Pendiente'],
                 pct_cols=['% Cartera / Facturacion'],
                 highlight_col='Riesgo Alto')

    wb.save(output_path)
